# -*- coding: utf-8 -*-
"""
Created on Mon Jun 20 15:11:33 2022

@author: jayden.qian
"""

import pandas as pd
import openpyxl 

#for BO Part 1
df1 = pd.read_excel("BO sheet template (most recent).xlsx",sheet_name ="BO page",header=0, usecols="B:H")
#for BO Part 2
df = pd.read_excel("BackOrderReport.xlsx",sheet_name ="Backorder detail",header=14, usecols="B:U")
#fill Nan company column 
df1['Company Name'].ffill(inplace=True)

def BO1(s):
    
    filtered_df1=df1.loc[(df1["Company Name"]==s) & (df1["Category"]=="Handset")]
    filtered_df=df.loc[(df["Customer"]==s) & (df["Category"]=="Handset"),["SRS ID","Date Created","Days of order","Customer Reference","Contact","User Name","Product","Qty"]]
      # write to more than one sheet in the workbook
    with pd.ExcelWriter("BO with list-python.xlsx",mode="a",engine="openpyxl",if_sheet_exists=("overlay")) as writer:  
          #for BO Part 1
          filtered_df1.to_excel(writer, index=False, sheet_name=s,startrow=0)
          #for BO Part 2
          filtered_df.to_excel(writer, index=False, sheet_name=s,startrow=10)
    
excel_file=openpyxl.load_workbook("BO with list-python.xlsx")
sheets=excel_file.sheetnames

for worksheet in excel_file.worksheets:
    worksheet.delete_rows(1,worksheet.max_row+1)
    
for sheet in sheets:
    
     BO1(sheet)
          
      
   
     
  



                   

                   
